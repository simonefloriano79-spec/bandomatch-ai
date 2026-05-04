"""
BandoMatch AI — Blueprint Enterprise
Gestisce il portafoglio clienti, la dashboard aggregata, l'export Excel
e gli alert prioritari per utenti con piano Enterprise.
"""
import io
import json
import logging
from datetime import datetime

from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, jsonify, send_file, abort
)
from flask_login import login_required, current_user

from extensions import db
from models.cliente_enterprise import ClienteEnterprise
from models.analisi import Analisi

logger = logging.getLogger(__name__)

enterprise_bp = Blueprint('enterprise', __name__, url_prefix='/enterprise')

PIANI_ENTERPRISE = {'enterprise'}


def _richiede_enterprise():
    """Verifica che l'utente abbia il piano Enterprise, altrimenti redirect."""
    piano = (current_user.piano or 'free').lower()
    if piano not in PIANI_ENTERPRISE:
        flash('Funzione riservata al piano Enterprise — Partner Istituzionali.', 'warning')
        return redirect(url_for('dashboard.home'))
    return None


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD AGGREGATA
# ─────────────────────────────────────────────────────────────────────────────

@enterprise_bp.route('/')
@enterprise_bp.route('/dashboard')
@login_required
def dashboard():
    """Dashboard aggregata: tutti i clienti e i loro match in un colpo."""
    guard = _richiede_enterprise()
    if guard:
        return guard

    clienti = (ClienteEnterprise.query
               .filter_by(partner_id=current_user.id, attivo=True)
               .order_by(ClienteEnterprise.data_aggiornamento.desc())
               .all())

    # KPI aggregati
    totale_clienti      = len(clienti)
    totale_bandi_verdi  = sum(c.bandi_verdi_ultimo  for c in clienti)
    totale_bandi_gialli = sum(c.bandi_gialli_ultimo for c in clienti)
    valore_totale       = sum(c.valore_potenziale   for c in clienti)

    # Clienti con alert prioritari (bandi verdi ad alto valore > €100K)
    clienti_alert = [c for c in clienti if c.valore_potenziale >= 100_000]

    return render_template(
        'enterprise_dashboard.html',
        clienti=clienti,
        totale_clienti=totale_clienti,
        totale_bandi_verdi=totale_bandi_verdi,
        totale_bandi_gialli=totale_bandi_gialli,
        valore_totale=valore_totale,
        clienti_alert=clienti_alert,
        nome_partner=current_user.nome_partner or current_user.email,
        logo_url=current_user.logo_url,
    )


# ─────────────────────────────────────────────────────────────────────────────
# GESTIONE CLIENTI
# ─────────────────────────────────────────────────────────────────────────────

@enterprise_bp.route('/clienti')
@login_required
def lista_clienti():
    """Lista completa clienti del portafoglio."""
    guard = _richiede_enterprise()
    if guard:
        return guard

    clienti = (ClienteEnterprise.query
               .filter_by(partner_id=current_user.id, attivo=True)
               .order_by(ClienteEnterprise.ragione_sociale)
               .all())
    return render_template('enterprise_clienti.html', clienti=clienti,
                           nome_partner=current_user.nome_partner or current_user.email)


@enterprise_bp.route('/clienti/nuovo', methods=['GET', 'POST'])
@login_required
def nuovo_cliente():
    """Aggiunge un nuovo cliente al portafoglio."""
    guard = _richiede_enterprise()
    if guard:
        return guard

    if request.method == 'POST':
        try:
            cliente = ClienteEnterprise(
                partner_id      = current_user.id,
                ragione_sociale = request.form.get('ragione_sociale', '').strip(),
                codice_fiscale  = request.form.get('codice_fiscale', '').strip() or None,
                partita_iva     = request.form.get('partita_iva', '').strip() or None,
                email_cliente   = request.form.get('email_cliente', '').strip() or None,
                ateco           = request.form.get('ateco', '').strip() or None,
                regione         = request.form.get('regione', '').strip() or None,
                forma_giuridica = request.form.get('forma_giuridica', '').strip() or None,
                note            = request.form.get('note', '').strip() or None,
                nome_partner    = current_user.nome_partner,
                logo_url        = current_user.logo_url,
            )
            if not cliente.ragione_sociale:
                flash('La ragione sociale è obbligatoria.', 'error')
                return render_template('enterprise_nuovo_cliente.html')

            db.session.add(cliente)
            db.session.commit()
            flash(f'Cliente "{cliente.ragione_sociale}" aggiunto al portafoglio.', 'success')
            return redirect(url_for('enterprise.lista_clienti'))
        except Exception as e:
            db.session.rollback()
            logger.error(f'Errore creazione cliente enterprise: {e}')
            flash('Errore durante il salvataggio. Riprova.', 'error')

    return render_template('enterprise_nuovo_cliente.html',
                           nome_partner=current_user.nome_partner or current_user.email)


@enterprise_bp.route('/clienti/<int:cliente_id>/elimina', methods=['POST'])
@login_required
def elimina_cliente(cliente_id: int):
    """Disattiva (soft-delete) un cliente dal portafoglio."""
    guard = _richiede_enterprise()
    if guard:
        return guard

    cliente = ClienteEnterprise.query.filter_by(
        id=cliente_id, utente_id=current_user.id
    ).first_or_404()
    cliente.attivo = False
    db.session.commit()
    flash(f'Cliente "{cliente.ragione_sociale}" rimosso dal portafoglio.', 'success')
    return redirect(url_for('enterprise.lista_clienti'))


@enterprise_bp.route('/clienti/<int:cliente_id>/aggiorna', methods=['POST'])
@login_required
def aggiorna_match_cliente(cliente_id: int):
    """
    Aggiorna i dati di match di un cliente recuperando l'ultima analisi
    associata alla sua ragione sociale.
    """
    guard = _richiede_enterprise()
    if guard:
        return guard

    cliente = ClienteEnterprise.query.filter_by(
        id=cliente_id, partner_id=current_user.id, attivo=True
    ).first_or_404()

    # Cerca l'analisi più recente per questo utente con la stessa ragione sociale
    analisi = (Analisi.query
               .filter_by(utente_id=current_user.id)
               .filter(Analisi.ragione_sociale.ilike(f'%{cliente.ragione_sociale[:20]}%'))
               .order_by(Analisi.data_analisi.desc())
               .first())

    if analisi:
        cliente.ultima_analisi_id   = analisi.id
        cliente.ultima_analisi_data = analisi.data_analisi
        cliente.bandi_verdi_ultimo  = analisi.bandi_verdi  or 0
        cliente.bandi_gialli_ultimo = analisi.bandi_gialli or 0
        cliente.valore_potenziale   = analisi.valore_potenziale or 0.0
        db.session.commit()
        return jsonify({'ok': True, 'analisi_id': analisi.id,
                        'bandi_verdi': cliente.bandi_verdi_ultimo,
                        'valore': cliente.valore_potenziale})
    return jsonify({'ok': False, 'msg': 'Nessuna analisi trovata per questo cliente.'})


# ─────────────────────────────────────────────────────────────────────────────
# DOSSIER WHITE-LABEL
# ─────────────────────────────────────────────────────────────────────────────

@enterprise_bp.route('/dossier/<int:analisi_id>')
@login_required
def dossier_whitelabel(analisi_id: int):
    """
    Genera il PDF Dossier con il logo del partner (white-label).
    Accessibile solo agli utenti Enterprise.
    """
    guard = _richiede_enterprise()
    if guard:
        return guard

    analisi = Analisi.query.filter_by(
        id=analisi_id, utente_id=current_user.id
    ).first_or_404()

    bandi_compatibili = []
    if analisi.risultati_json:
        try:
            dati = json.loads(analisi.risultati_json)
            bandi_raw = dati.get('risultati', dati.get('bandi', [])) if isinstance(dati, dict) else dati
            bandi_compatibili = [b for b in bandi_raw
                                 if str(b.get('semaforo', '')).upper() in ('VERDE', 'GIALLO')]
        except (json.JSONDecodeError, TypeError):
            pass

    from utils.dossier import genera_dossier
    pdf_bytes = genera_dossier(
        current_user,
        analisi,
        bandi_compatibili,
        logo_url=current_user.logo_url,
        nome_partner=current_user.nome_partner,
    )

    nome_file = (f"dossier_{analisi.ragione_sociale or 'azienda'}"
                 f"_{datetime.now().strftime('%Y%m%d')}.pdf"
                 ).replace(' ', '_')

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=nome_file,
    )


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL MASSIVO
# ─────────────────────────────────────────────────────────────────────────────

@enterprise_bp.route('/export/excel')
@login_required
def export_excel():
    """
    Esporta tutti i match di tutti i clienti in un file Excel.
    Ogni cliente occupa un foglio separato.
    """
    guard = _richiede_enterprise()
    if guard:
        return guard

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('Libreria openpyxl non disponibile. Contatta il supporto.', 'error')
        return redirect(url_for('enterprise.dashboard'))

    clienti = (ClienteEnterprise.query
               .filter_by(partner_id=current_user.id, attivo=True)
               .order_by(ClienteEnterprise.ragione_sociale)
               .all())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuove foglio vuoto di default

    # ── Foglio riepilogo ──
    ws_riepilogo = wb.create_sheet('Riepilogo Portafoglio')
    header_fill  = PatternFill('solid', fgColor='0F172A')
    header_font  = Font(color='FFFFFF', bold=True)
    verde_fill   = PatternFill('solid', fgColor='DCFCE7')
    giallo_fill  = PatternFill('solid', fgColor='FEF9C3')

    riepilogo_headers = [
        'Ragione Sociale', 'ATECO', 'Regione', 'Forma Giuridica',
        'Bandi Verdi', 'Bandi Gialli', 'Valore Potenziale (€)',
        'Ultima Analisi', 'Email Cliente'
    ]
    ws_riepilogo.append(riepilogo_headers)
    for col_idx, _ in enumerate(riepilogo_headers, 1):
        cell = ws_riepilogo.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for cliente in clienti:
        row = [
            cliente.ragione_sociale,
            cliente.ateco or '',
            cliente.regione or '',
            cliente.forma_giuridica or '',
            cliente.bandi_verdi_ultimo,
            cliente.bandi_gialli_ultimo,
            round(cliente.valore_potenziale or 0, 2),
            cliente.ultima_analisi_data.strftime('%d/%m/%Y') if cliente.ultima_analisi_data else '',
            cliente.email_cliente or '',
        ]
        ws_riepilogo.append(row)
        r = ws_riepilogo.max_row
        if cliente.bandi_verdi_ultimo > 0:
            ws_riepilogo.cell(row=r, column=5).fill = verde_fill
        if cliente.bandi_gialli_ultimo > 0:
            ws_riepilogo.cell(row=r, column=6).fill = giallo_fill

    # Auto-width colonne riepilogo
    for col in ws_riepilogo.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws_riepilogo.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Un foglio per ogni cliente con i bandi dettagliati ──
    for cliente in clienti:
        if not cliente.ultima_analisi_id:
            continue
        analisi = Analisi.query.get(cliente.ultima_analisi_id)
        if not analisi or not analisi.risultati_json:
            continue

        try:
            dati = json.loads(analisi.risultati_json)
            bandi_raw = dati.get('risultati', dati.get('bandi', [])) if isinstance(dati, dict) else dati
        except (json.JSONDecodeError, TypeError):
            continue

        nome_foglio = cliente.ragione_sociale[:28].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(nome_foglio)

        bando_headers = [
            'Bando', 'Ente', 'Semaforo', 'Score (%)',
            'Importo Max (€)', '% Fondo Perduto', 'Scadenza', 'URL'
        ]
        ws.append(bando_headers)
        for col_idx, _ in enumerate(bando_headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for b in bandi_raw:
            sem = str(b.get('semaforo', '')).upper()
            agev = b.get('agevolazioni', {})
            massimale = (b.get('massimale_agevolazione') or
                         agev.get('massimale_investimento') or
                         agev.get('spesa_progetto_max') or '')
            perc_fp = (b.get('percentuale_fondo_perduto') or
                       agev.get('percentuale_fondo_perduto') or
                       agev.get('percentuale_fondo_perduto_fino_120k') or '')
            scadenza = b.get('data_scadenza') or b.get('stato_bando') or ''
            ws.append([
                b.get('titolo', b.get('bando_nome', '')),
                b.get('fonte', b.get('ente', '')),
                sem,
                b.get('score') or '',
                massimale,
                f"{perc_fp}%" if perc_fp else '',
                str(scadenza)[:10],
                b.get('url', ''),
            ])
            r = ws.max_row
            if sem == 'VERDE':
                for c in range(1, 9):
                    ws.cell(row=r, column=c).fill = verde_fill
            elif sem == 'GIALLO':
                for c in range(1, 9):
                    ws.cell(row=r, column=c).fill = giallo_fill

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Salva in buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_file = (f"BandoMatch_Export_{current_user.nome_partner or 'partner'}"
                 f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                 ).replace(' ', '_')

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_file,
    )


# ─────────────────────────────────────────────────────────────────────────────
# ALERT PRIORITARI (API JSON)
# ─────────────────────────────────────────────────────────────────────────────

@enterprise_bp.route('/alert/prioritari')
@login_required
def alert_prioritari():
    """
    Restituisce i clienti con bandi ad alto valore (>€100K) come JSON.
    Usato dalla dashboard per il pannello alert.
    """
    guard = _richiede_enterprise()
    if guard:
        return jsonify({'error': 'Enterprise only'}), 403

    clienti_alert = (ClienteEnterprise.query
                     .filter_by(partner_id=current_user.id, attivo=True)
                     .filter(ClienteEnterprise.valore_potenziale >= 100_000)
                     .order_by(ClienteEnterprise.valore_potenziale.desc())
                     .all())

    return jsonify({
        'alert': [
            {
                'cliente':          c.ragione_sociale,
                'valore':           c.valore_potenziale,
                'bandi_verdi':      c.bandi_verdi_ultimo,
                'ultima_analisi':   c.ultima_analisi_data.isoformat() if c.ultima_analisi_data else None,
                'analisi_id':       c.ultima_analisi_id,
            }
            for c in clienti_alert
        ],
        'totale': len(clienti_alert),
    })


# ─────────────────────────────────────────────────────────────────────────────
# IMPOSTAZIONI PARTNER (white-label)
# ─────────────────────────────────────────────────────────────────────────────

@enterprise_bp.route('/impostazioni', methods=['GET', 'POST'])
@login_required
def impostazioni():
    """Configura il branding white-label del partner (nome e logo)."""
    guard = _richiede_enterprise()
    if guard:
        return guard

    if request.method == 'POST':
        current_user.nome_partner = request.form.get('nome_partner', '').strip() or None
        current_user.logo_url     = request.form.get('logo_url', '').strip() or None
        db.session.commit()
        flash('Impostazioni partner aggiornate.', 'success')
        return redirect(url_for('enterprise.impostazioni'))

    return render_template('enterprise_impostazioni.html',
                           nome_partner=current_user.nome_partner or '',
                           logo_url=current_user.logo_url or '')
